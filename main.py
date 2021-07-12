import shutil

import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime
import os
import sys
import os.path
from os import path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Border, Side


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    pass


def companyNameLookUp(companyName):
    companyDict = {
        'ALL': 'All Seasons Realty Corp.',
        'APL': 'Allianz-PNB Life Insurance, Inc. (APLII)',
        'ABI': 'Asia Brewery, Inc. (ABI), Subsidiaries',
        'BHC': 'Basic Holdings Corp.',
        'CPH': 'Century Park Hotel',
        # "EPP": "Eton Properties Philippines, Inc. (EPPI), Subsidiaries",
        "EPP": "Eton Properties Philippines, Inc. (EPPI), Subsidiaries",
        'FFI': 'Foremost Farms, Inc.',
        'FTC': 'Fortune Tobacco Corp.',
        'GDC': 'Grandspan Development Corp.',
        'HII': 'Himmel Industries, Inc.',
        'LRC': 'Landcom Realty Corp.',
        # 'LRC': 'LANDCOM REALTY CORP.',
        'LTG': 'LT Group, Inc. (Parent Company)',
        'DIR': 'LTGC Directors',
        'MAC': 'MacroAsia Corp., Subsidiaries and Affiliates',
        'PAL': 'Philippine Airlines, Inc. (PAL), Subsidiaries and Affiliates',
        'PNB': 'Philippine National Bank (PNB), Subsidiaries',
        'PMI': 'PMFTC',
        'RAP': 'Rapid Movers & Forwarders, Inc.',
        'TYK': 'Tan Yan Kee Foundation, Inc. (TYKFI)',
        'TDI': 'Tanduay Distillers, Inc. (TDI), Subsidiaries',
        'CHI': 'Charter House Inc.',
        'SPV': 'SPV-AMC Group',
        'TMC': 'Topkick Movers Corporation',
        'UNI': 'University of the East (UE)',
        'UER': 'University of the East Ramon Magsaysay Memorial Medical Center (UERMMMC)',
        'VMC': 'Victorias Milling Company, Inc. (VMC)',
        'ZHI': 'Zebra Holdings, Inc.',
        'STN': 'Sabre Travel Network Phils., Inc.',
        'PAN': 'Pan Asia Securities',
        'ANA': 'All Nippon Airways',
        'LTC': 'LUCKY TRAVEL CORPORATION',
    }
    company_Code = ""
    for key, value in companyDict.items():
        if companyName.strip() == value:
            company_Code = key

    return company_Code


def duplicateTemplateHHLTGC(tempLTGC_Path, out, compCode, companyName):
    companyDir = out

    # # creating new DIR base on company code
    # if not path.exists(out + "/" + compCode):
    #     os.mkdir(os.path.join(out, compCode))

    # # creating new DIR base on company code
    # os.mkdir(os.path.join(out, compCode))

    # shutil.copy(tempLTGC_Path,
    #             companyDir + "/" + companyName + "_HH.xlsx")
    #
    # return companyDir + "/" + companyName + "_HH.xlsx"

    shutil.copy(tempLTGC_Path,
                companyDir + "/" + companyName + "_HHLTGC_CEIRMasterlist.xlsx")

    return companyDir + "/" + companyName + "_HHLTGC_CEIRMasterlist.xlsx"


def addingDataValidation(currentSheet, numrows):
    Category_data_val = DataValidation(type="list", formula1="=LOVCategories")
    currentSheet.add_data_validation(Category_data_val)

    CategoryID_data_val = DataValidation(type="list", formula1="=LOVCategoryID")
    currentSheet.add_data_validation(CategoryID_data_val)

    Suffix_data_val = DataValidation(type="list", formula1="=LOVSuffix")
    currentSheet.add_data_validation(Suffix_data_val)

    C_residence_region_data_val = DataValidation(type="list", formula1="=Region")
    currentSheet.add_data_validation(C_residence_region_data_val)

    C_residence_province_data_val = DataValidation(type="list", formula1="=INDIRECT(L3)")
    currentSheet.add_data_validation(C_residence_province_data_val)

    C_residence_municipality_data_val = DataValidation(type="list", formula1="=INDIRECT(M3)")
    currentSheet.add_data_validation(C_residence_municipality_data_val)

    C_residence_Barangay_data_val = DataValidation(type="list", formula1="=INDIRECT(N3)")
    currentSheet.add_data_validation(C_residence_Barangay_data_val)

    sex_data_val = DataValidation(type="list", formula1="=LOVSex")
    currentSheet.add_data_validation(sex_data_val)

    civilStatus_data_val = DataValidation(type="list", formula1="=LOVCivilStatus")
    currentSheet.add_data_validation(civilStatus_data_val)

    employmentStatus_data_val = DataValidation(type="list", formula1="=LOVEmploymentStatus")
    currentSheet.add_data_validation(employmentStatus_data_val)

    Directly_in_interaction_with_COVID_patient_data_val = DataValidation(type="list", formula1="=LOVDirectCovid")
    currentSheet.add_data_validation(Directly_in_interaction_with_COVID_patient_data_val)

    Profession_data_val = DataValidation(type="list", formula1="=LOVProfession")
    currentSheet.add_data_validation(Profession_data_val)

    ICC_of_Employer_data_val = DataValidation(type="list", formula1="=LOVProvinceHUCICCofEmployer")
    currentSheet.add_data_validation(ICC_of_Employer_data_val)

    Pregnancy_status_data_val = DataValidation(type="list", formula1="=LOVPregnancyStatus")
    currentSheet.add_data_validation(Pregnancy_status_data_val)

    YesNo_data_val = DataValidation(type="list", formula1="=LOVYesNo")
    currentSheet.add_data_validation(YesNo_data_val)

    With_Comorbidity_data_val = DataValidation(type="list", formula1="=LOVYesNone")
    currentSheet.add_data_validation(With_Comorbidity_data_val)

    Classification_of_COVID_19_data_val = DataValidation(type="list", formula1="=LOVCovidClass")
    currentSheet.add_data_validation(Classification_of_COVID_19_data_val)

    Willing_to_be_Vaccinated_data_val = DataValidation(type="list", formula1="=LOVConsent")
    currentSheet.add_data_validation(Willing_to_be_Vaccinated_data_val)

    Signup_coompletion_Time_data_val = DataValidation(type="list", formula1="=LOVWFH")
    currentSheet.add_data_validation(Signup_coompletion_Time_data_val)

    A2_Senior_data_val = DataValidation(type="list", formula1="=A2LOV")
    currentSheet.add_data_validation(A2_Senior_data_val)

    A3_With_Co_morbidity_data_val = DataValidation(type="list", formula1="=A3LOV")
    currentSheet.add_data_validation(A3_With_Co_morbidity_data_val)

    AgeRiskFactor_data_val = DataValidation(type="list", formula1="=AgeRiskFactor")  # 55-59_y/o
    currentSheet.add_data_validation(AgeRiskFactor_data_val)

    Confirmed_Vaccination_Site_data_val = DataValidation(type="list", formula1="=VaccinationSites")
    currentSheet.add_data_validation(Confirmed_Vaccination_Site_data_val)

    row = numrows + 3
    Category_data_val.add("A3:A" + str(row))
    CategoryID_data_val.add("B3:B" + str(row))
    Suffix_data_val.add("I3:I" + str(row))
    C_residence_region_data_val.add("L3:L" + str(row))
    C_residence_province_data_val.add("M3:M" + str(row))
    C_residence_municipality_data_val.add("N3:N" + str(row))
    C_residence_Barangay_data_val.add("O3:O" + str(row))
    sex_data_val.add("P3:P" + str(row))
    civilStatus_data_val.add("R3:R" + str(row))
    employmentStatus_data_val.add("S3:S" + str(row))
    Directly_in_interaction_with_COVID_patient_data_val.add("T3:T" + str(row))
    Profession_data_val.add("U3:U" + str(row))
    ICC_of_Employer_data_val.add("W3:W" + str(row))
    Pregnancy_status_data_val.add("Z3:Z" + str(row))
    YesNo_data_val.add("AA3:AA" + str(row))
    YesNo_data_val.add("AB3:AB" + str(row))
    YesNo_data_val.add("AC3:AC" + str(row))
    YesNo_data_val.add("AD3:AD" + str(row))
    YesNo_data_val.add("AE3:AE" + str(row))
    YesNo_data_val.add("AF3:AF" + str(row))
    YesNo_data_val.add("AG3:AG" + str(row))
    With_Comorbidity_data_val.add("AH3:AH" + str(row))
    YesNo_data_val.add("AI3:AI" + str(row))
    YesNo_data_val.add("AJ3:AJ" + str(row))
    YesNo_data_val.add("AK3:AK" + str(row))
    YesNo_data_val.add("AL3:AL" + str(row))
    YesNo_data_val.add("AM3:AM" + str(row))
    YesNo_data_val.add("AN3:AN" + str(row))
    YesNo_data_val.add("AO3:AO" + str(row))
    YesNo_data_val.add("AP3:AP" + str(row))
    YesNo_data_val.add("AQ3:AQ" + str(row))
    Classification_of_COVID_19_data_val.add("AS3:AS" + str(row))
    Willing_to_be_Vaccinated_data_val.add("AT3:AT" + str(row))
    A2_Senior_data_val.add("BD3:BD" + str(row))
    A3_With_Co_morbidity_data_val.add("BE3:BE" + str(row))
    # AgeRiskFactor_data_val.add("BF3:BF" + str(row))
    Confirmed_Vaccination_Site_data_val.add("BL3:BL" + str(row))

    # # set data validation(dropdown)
    # for r in range(4, numrows + 3):
    # Category_data_val.add(currentSheet["A" + str(r)])
    # CategoryID_data_val.add(currentSheet["B" + str(r)])
    # Suffix_data_val.add(currentSheet["I" + str(r)])
    # C_residence_region_data_val.add(currentSheet["L" + str(r)])
    # C_residence_province_data_val.add(currentSheet["M" + str(r)])
    # C_residence_municipality_data_val.add(currentSheet["N" + str(r)])
    # C_residence_Barangay_data_val.add(currentSheet["O" + str(r)])
    # sex_data_val.add(currentSheet["P" + str(r)])
    # civilStatus_data_val.add(currentSheet["R" + str(r)])
    # employmentStatus_data_val.add(currentSheet["S" + str(r)])
    # Directly_in_interaction_with_COVID_patient_data_val.add(currentSheet["T" + str(r)])
    # Profession_data_val.add(currentSheet["U" + str(r)])
    # ICC_of_Employer_data_val.add(currentSheet["W" + str(r)])
    # Pregnancy_status_data_val.add(currentSheet["Z" + str(r)])
    # YesNo_data_val.add(currentSheet["AA" + str(r)])
    # YesNo_data_val.add(currentSheet["AB" + str(r)])
    # YesNo_data_val.add(currentSheet["AC" + str(r)])
    # YesNo_data_val.add(currentSheet["AD" + str(r)])
    # YesNo_data_val.add(currentSheet["AE" + str(r)])
    # YesNo_data_val.add(currentSheet["AF" + str(r)])
    # YesNo_data_val.add(currentSheet["AG" + str(r)])
    # With_Comorbidity_data_val.add((currentSheet["AH" + str(r)]))
    # YesNo_data_val.add(currentSheet["AI" + str(r)])
    # YesNo_data_val.add(currentSheet["AJ" + str(r)])
    # YesNo_data_val.add(currentSheet["AK" + str(r)])
    # YesNo_data_val.add(currentSheet["AL" + str(r)])
    # YesNo_data_val.add(currentSheet["AM" + str(r)])
    # YesNo_data_val.add(currentSheet["AN" + str(r)])
    # YesNo_data_val.add(currentSheet["AO" + str(r)])
    # YesNo_data_val.add(currentSheet["AP" + str(r)])
    # YesNo_data_val.add(currentSheet["AQ" + str(r)])
    # Classification_of_COVID_19_data_val.add(currentSheet["AS" + str(r)])
    # Willing_to_be_Vaccinated_data_val.add(currentSheet["AT" + str(r)])
    # # Signup_coompletion_Time_data_val.add(currentSheet["AY"+str(r)])
    # A2_Senior_data_val.add(currentSheet["BD" + str(r)])
    # A3_With_Co_morbidity_data_val.add(currentSheet["BE" + str(r)])
    # AgeRiskFactor_data_val.add(currentSheet["BF" + str(r)])
    # Confirmed_Vaccination_Site_data_val.add(currentSheet["BL" + str(r)])

    pass


def getData(inFile_HHLTGC, outPath):
    print("============================")
    print("Starting HH LTGC Files")
    print("============================")

    df = pd.read_excel(inFile_HHLTGC, sheet_name='Eligible Population', header=1,
                       dtype={'PhilHealth_ID*': str, 'Contact_number_of_employer*': str,
                              'Contact_No.*': str, 'Age': str}, na_filter=False)

    groups = df.groupby('Company Name')

    for i, comp in groups:
        companyCode = companyNameLookUp(i)

        comp = comp.astype(str)

        # print(comp)

        # get num rows
        numrows = len(comp.index)

        print(i + ' (' + companyCode + ") has " + str(numrows) + " records", end='')

        templateFile = duplicateTemplateHHLTGC(tempHHLTGC_Path, outPath, companyCode, i)

        # border settings
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        theFile = openpyxl.load_workbook(templateFile)
        currentSheet = theFile["Eligible Population"]
        addingDataValidation(currentSheet, numrows)

        # set_border(currentSheet, "A3:BN"+str(numrows+2))
        set_border(currentSheet, "A3:BL"+str(numrows+2))

        theFile.save(templateFile)

        writer = pd.ExcelWriter(templateFile, engine='openpyxl', mode='a')
        writer.book = load_workbook(templateFile)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        comp.to_excel(writer, "Eligible Population", startrow=2, header=False, index=False)
        writer.save()

        print(".....Done!")
    pass


if __name__ == '__main__':
    pd.set_option('display.max_columns', None)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.width', None)

    # get files in "in" folder

    today = datetime.today()
    dateTime = today.strftime("%m_%d_%y_%H%M%S")

    inPath = r"/Users/Ran/Documents/Vaccine/LTGSplit/in"
    outPath = r"/Users/Ran/Documents/Vaccine/LTGSplit/out/hhltgc"
    templateFilePath = r"/Users/Ran/Documents/Vaccine/LTGSplit/template"

    # inFile_HHLTGC = inPath + "/HHLTGC_CEIRMasterlist.xlsx"
    inFile_HHLTGC = inPath + "/HHLTGC_CEIRMasterlist_0625_1038AM.xlsx"

    tempHHLTGC_Path = templateFilePath + "/HHLTGC_CEIRMasterlist_ExtraCols.xlsx"

    # Excel Templates: create copy

    print("Split File Script......")

    if path.exists(inFile_HHLTGC) and path.isfile(inFile_HHLTGC):
        getData(inFile_HHLTGC, outPath)
    else:
        print(str(inFile_HHLTGC) + " File is invalid or does not exist")
